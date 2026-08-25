from __future__ import annotations

import csv
import hashlib
import io
import re
import uuid
from datetime import datetime, timezone
from decimal import Decimal, InvalidOperation
from typing import Any

from openpyxl import load_workbook
from pydantic import BaseModel, ConfigDict, Field
from sqlalchemy import DateTime, Index, JSON, Numeric, String, Text, UniqueConstraint, Uuid, delete, select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import Mapped, mapped_column

from app.core.database import Base
from app.core.errors import AppError


REQUIRED_COLUMNS = {
    "id",
    "contractname",
    "enteredat",
    "exitedat",
    "entryprice",
    "exitprice",
    "size",
    "type",
}


class Trade(Base):
    __tablename__ = "trades"
    __table_args__ = (
        UniqueConstraint("source_file_hash", "source_trade_id", name="uq_trade_source"),
        Index("ix_trades_time_window", "entered_at", "exited_at"),
    )

    id: Mapped[uuid.UUID] = mapped_column(Uuid(as_uuid=True), primary_key=True, default=uuid.uuid4)
    source_trade_id: Mapped[str] = mapped_column(String(120), nullable=False)
    contract_name: Mapped[str] = mapped_column(String(40), nullable=False)
    symbol_root: Mapped[str] = mapped_column(String(20), nullable=False, index=True)
    entered_at: Mapped[datetime] = mapped_column(DateTime(timezone=True), nullable=False)
    exited_at: Mapped[datetime] = mapped_column(DateTime(timezone=True), nullable=False)
    entry_price: Mapped[Decimal] = mapped_column(Numeric(20, 8), nullable=False)
    exit_price: Mapped[Decimal] = mapped_column(Numeric(20, 8), nullable=False)
    direction: Mapped[str] = mapped_column(String(10), nullable=False)
    size: Mapped[Decimal] = mapped_column(Numeric(20, 8), nullable=False)
    reported_pnl: Mapped[Decimal | None] = mapped_column(Numeric(20, 8))
    fees: Mapped[Decimal | None] = mapped_column(Numeric(20, 8))
    commissions: Mapped[Decimal | None] = mapped_column(Numeric(20, 8))
    trade_day: Mapped[datetime | None] = mapped_column(DateTime(timezone=True))
    trade_duration_seconds: Mapped[Decimal | None] = mapped_column(Numeric(20, 8))
    source_file_name: Mapped[str] = mapped_column(String(255), nullable=False)
    source_file_hash: Mapped[str] = mapped_column(String(64), nullable=False)
    imported_at: Mapped[datetime] = mapped_column(DateTime(timezone=True), default=lambda: datetime.now(timezone.utc))
    raw_data: Mapped[dict[str, Any]] = mapped_column(JSON, nullable=False)
    slippage: Mapped[Decimal | None] = mapped_column(Numeric(20, 8))
    strategy: Mapped[str | None] = mapped_column(String(120))
    account: Mapped[str | None] = mapped_column(String(120))
    tags: Mapped[list[str]] = mapped_column(JSON, default=list)
    notes: Mapped[str] = mapped_column(Text, default="")
    attachments: Mapped[list[str]] = mapped_column(JSON, default=list)


class TradeImportBatch(Base):
    __tablename__ = "trade_import_batches"

    id: Mapped[uuid.UUID] = mapped_column(Uuid(as_uuid=True), primary_key=True, default=uuid.uuid4)
    file_name: Mapped[str] = mapped_column(String(255))
    file_hash: Mapped[str] = mapped_column(String(64), index=True)
    total_rows: Mapped[int] = mapped_column()
    valid_rows: Mapped[int] = mapped_column()
    invalid_rows: Mapped[int] = mapped_column()
    imported_rows: Mapped[int] = mapped_column(default=0)
    skipped_duplicates: Mapped[int] = mapped_column(default=0)
    mapping: Mapped[dict[str, Any]] = mapped_column(JSON, default=dict)
    errors: Mapped[list[dict[str, Any]]] = mapped_column(JSON, default=list)
    created_at: Mapped[datetime] = mapped_column(DateTime(timezone=True), default=lambda: datetime.now(timezone.utc), index=True)


class TradeRow(BaseModel):
    model_config = ConfigDict(from_attributes=True)

    source_trade_id: str
    contract_name: str
    symbol_root: str
    entered_at: datetime
    exited_at: datetime
    entry_price: Decimal
    exit_price: Decimal
    direction: str
    size: Decimal
    reported_pnl: Decimal | None = None
    fees: Decimal | None = None
    commissions: Decimal | None = None
    trade_day: datetime | None = None
    trade_duration_seconds: Decimal | None = None
    raw_data: dict[str, Any]


class TradePreview(BaseModel):
    file_name: str
    file_hash: str
    total_rows: int
    valid_rows: int
    invalid_rows: int
    rows: list[TradeRow]
    errors: list[dict[str, Any]]


class TradeImportResult(BaseModel):
    imported: int
    skipped_duplicates: int
    total: int


class TradeResponse(TradeRow):
    id: uuid.UUID
    source_file_name: str
    imported_at: datetime
    slippage: Decimal | None = None
    strategy: str | None = None
    account: str | None = None
    tags: list[str] = Field(default_factory=list)
    notes: str = ""
    attachments: list[str] = Field(default_factory=list)


class TradeCreate(BaseModel):
    contract_name: str
    entered_at: datetime
    exited_at: datetime
    entry_price: Decimal
    exit_price: Decimal
    direction: str
    size: Decimal
    reported_pnl: Decimal | None = None
    fees: Decimal | None = None
    commissions: Decimal | None = None
    slippage: Decimal | None = None
    strategy: str | None = Field(None, max_length=120)
    account: str | None = Field(None, max_length=120)
    tags: list[str] = Field(default_factory=list)
    notes: str = Field("", max_length=10000)
    attachments: list[str] = Field(default_factory=list)


class TradeUpdate(BaseModel):
    entered_at: datetime | None = None
    exited_at: datetime | None = None
    entry_price: Decimal | None = None
    exit_price: Decimal | None = None
    direction: str | None = None
    size: Decimal | None = None
    reported_pnl: Decimal | None = None
    fees: Decimal | None = None
    commissions: Decimal | None = None
    slippage: Decimal | None = None
    strategy: str | None = Field(None, max_length=120)
    account: str | None = Field(None, max_length=120)
    tags: list[str] | None = None
    notes: str | None = Field(None, max_length=10000)
    attachments: list[str] | None = None


class TradeImportBatchResponse(BaseModel):
    id: uuid.UUID
    file_name: str
    file_hash: str
    total_rows: int
    valid_rows: int
    invalid_rows: int
    imported_rows: int
    skipped_duplicates: int
    mapping: dict[str, Any]
    errors: list[dict[str, Any]]
    created_at: datetime
    model_config = ConfigDict(from_attributes=True)


def _normalize_header(value: Any) -> str:
    return re.sub(r"[^a-z0-9]", "", str(value or "").strip().lower())


def _parse_decimal(value: Any, field: str) -> Decimal:
    try:
        return Decimal(str(value).replace(",", "").strip())
    except (InvalidOperation, AttributeError):
        raise ValueError(f"{field} 不是有效数字") from None


def _optional_decimal(value: Any) -> Decimal | None:
    if value is None or str(value).strip() == "":
        return None
    return _parse_decimal(value, "数值")


def _parse_datetime(value: Any, field: str) -> datetime:
    if isinstance(value, datetime):
        parsed = value
    else:
        text = str(value or "").strip().replace("Z", "+00:00")
        try:
            parsed = datetime.fromisoformat(text)
        except ValueError:
            try:
                parsed = datetime.strptime(text, "%m/%d/%Y %H:%M:%S %z")
            except ValueError:
                raise ValueError(f"{field} 不是有效时间") from None
    if parsed.tzinfo is None:
        parsed = parsed.replace(tzinfo=timezone.utc)
    return parsed.astimezone(timezone.utc)


def _optional_datetime(value: Any) -> datetime | None:
    if value is None or str(value).strip() == "":
        return None
    return _parse_datetime(value, "TradeDay")


def _optional_duration_seconds(value: Any) -> Decimal | None:
    if value is None or str(value).strip() == "":
        return None
    text = str(value).strip()
    if ":" not in text:
        return _parse_decimal(text, "TradeDuration")
    try:
        hours, minutes, seconds = text.split(":")
        return Decimal(hours) * 3600 + Decimal(minutes) * 60 + Decimal(seconds)
    except (ValueError, InvalidOperation):
        raise ValueError("TradeDuration 不是有效时长") from None


def _symbol_root(contract: str) -> str:
    match = re.match(r"^([A-Z]+?)(?:[FGHJKMNQUVXZ]\d{1,2})?$", contract.upper())
    return match.group(1) if match else contract.upper()


def _direction(value: Any) -> str:
    normalized = str(value or "").strip().lower()
    if normalized in {"long", "buy", "多", "多单"}:
        return "long"
    if normalized in {"short", "sell", "空", "空单"}:
        return "short"
    raise ValueError("Type 必须是 Long 或 Short")


def _read_rows(file_name: str, content: bytes) -> list[dict[str, Any]]:
    lower_name = file_name.lower()
    if content.startswith(b"PK") and b"Index/Document.iwa" in content:
        raise AppError("numbers_file_unsupported", "检测到 Apple Numbers 文件，请先导出为 .xlsx 或真正的 .csv 文件", 422)
    if lower_name.endswith(".xlsx"):
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        sheet = workbook.active
        values = list(sheet.iter_rows(values_only=True))
        if not values:
            return []
        headers = [str(value or "") for value in values[0]]
        return [dict(zip(headers, row, strict=False)) for row in values[1:] if any(value is not None for value in row)]
    if lower_name.endswith(".csv"):
        try:
            text = content.decode("utf-8-sig")
        except UnicodeDecodeError:
            text = content.decode("gb18030")
        return list(csv.DictReader(io.StringIO(text)))
    raise AppError("unsupported_trade_file", "仅支持 .xlsx 和 .csv 交易文件", 422)


def parse_trade_file(file_name: str, content: bytes) -> TradePreview:
    rows = _read_rows(file_name, content)
    file_hash = hashlib.sha256(content).hexdigest()
    if not rows:
        raise AppError("empty_trade_file", "交易文件没有数据行", 422)
    normalized_rows = [{_normalize_header(key): value for key, value in row.items()} for row in rows]
    missing = sorted(REQUIRED_COLUMNS - set(normalized_rows[0]))
    if missing:
        raise AppError("missing_trade_columns", "交易文件缺少必需字段", 422, [{"columns": missing}])

    valid: list[TradeRow] = []
    errors: list[dict[str, Any]] = []
    for index, row in enumerate(normalized_rows, start=2):
        try:
            entered_at = _parse_datetime(row.get("enteredat"), "EnteredAt")
            exited_at = _parse_datetime(row.get("exitedat"), "ExitedAt")
            if exited_at < entered_at:
                raise ValueError("ExitedAt 不能早于 EnteredAt")
            contract_name = str(row.get("contractname") or "").strip().upper()
            valid.append(TradeRow(
                source_trade_id=str(row.get("id") or "").strip(),
                contract_name=contract_name,
                symbol_root=_symbol_root(contract_name),
                entered_at=entered_at,
                exited_at=exited_at,
                entry_price=_parse_decimal(row.get("entryprice"), "EntryPrice"),
                exit_price=_parse_decimal(row.get("exitprice"), "ExitPrice"),
                direction=_direction(row.get("type")),
                size=_parse_decimal(row.get("size"), "Size"),
                reported_pnl=_optional_decimal(row.get("pnl")),
                fees=_optional_decimal(row.get("fees")),
                commissions=_optional_decimal(row.get("commissions")),
                trade_day=_optional_datetime(row.get("tradeday")),
                trade_duration_seconds=_optional_duration_seconds(row.get("tradeduration")),
                raw_data={key: "" if value is None else str(value) for key, value in row.items()},
            ))
        except ValueError as exc:
            errors.append({"row": index, "message": str(exc)})
    return TradePreview(
        file_name=file_name,
        file_hash=file_hash,
        total_rows=len(rows),
        valid_rows=len(valid),
        invalid_rows=len(errors),
        rows=valid,
        errors=errors[:100],
    )


async def import_trades(session: AsyncSession, file_name: str, content: bytes) -> TradeImportResult:
    preview = parse_trade_file(file_name, content)
    rows = _read_rows(file_name, content)
    normalized_rows = [{_normalize_header(key): value for key, value in row.items()} for row in rows]
    parsed = parse_trade_file(file_name, content)
    valid_by_id = {row.source_trade_id: row for row in parsed.rows}
    imported = 0
    skipped = 0
    for raw in normalized_rows:
        source_id = str(raw.get("id") or "").strip()
        row = valid_by_id.get(source_id)
        if row is None:
            continue
        exists = await session.scalar(select(Trade.id).where(
            Trade.source_file_hash == preview.file_hash,
            Trade.source_trade_id == source_id,
        ))
        if exists:
            skipped += 1
            continue
        session.add(Trade(
            **row.model_dump(exclude={"raw_data"}),
            raw_data=row.raw_data,
            source_file_name=file_name,
            source_file_hash=preview.file_hash,
        ))
        imported += 1
    await session.commit()
    session.add(TradeImportBatch(file_name=file_name, file_hash=preview.file_hash, total_rows=preview.total_rows, valid_rows=preview.valid_rows, invalid_rows=preview.invalid_rows, imported_rows=imported, skipped_duplicates=skipped, mapping={"headers": list(normalized_rows[0]) if normalized_rows else []}, errors=preview.errors))
    await session.commit()
    return TradeImportResult(imported=imported, skipped_duplicates=skipped, total=preview.valid_rows)


async def query_trades(
    session: AsyncSession,
    start: datetime,
    end: datetime,
    symbol: str | None = None,
) -> list[Trade]:
    statement = select(Trade).where(Trade.entered_at <= end, Trade.exited_at >= start)
    if symbol:
        statement = statement.where(Trade.symbol_root == symbol.upper())
    result = await session.scalars(
        statement.order_by(Trade.entered_at)
    )
    return list(result)


async def list_recent_trades(session: AsyncSession, limit: int = 200) -> list[Trade]:
    result = await session.scalars(
        select(Trade).order_by(Trade.entered_at.desc()).limit(limit)
    )
    return list(result)


async def create_trade(session: AsyncSession, payload: TradeCreate) -> Trade:
    if payload.exited_at < payload.entered_at:
        raise AppError("invalid_trade_time", "平仓时间不能早于开仓时间", 422)
    contract = payload.contract_name.strip().upper()
    trade = Trade(source_trade_id=f"manual-{uuid.uuid4()}", contract_name=contract, symbol_root=_symbol_root(contract), source_file_name="manual", source_file_hash="manual", raw_data={}, **payload.model_dump(exclude={"contract_name"}))
    session.add(trade)
    await session.commit(); await session.refresh(trade)
    return trade


async def update_trade(session: AsyncSession, trade_id: uuid.UUID, payload: TradeUpdate) -> Trade:
    trade = await session.get(Trade, trade_id)
    if trade is None:
        raise AppError("trade_not_found", "交易记录不存在", 404)
    for key, value in payload.model_dump(exclude_unset=True).items():
        setattr(trade, key, value)
    if trade.exited_at < trade.entered_at:
        raise AppError("invalid_trade_time", "平仓时间不能早于开仓时间", 422)
    await session.commit(); await session.refresh(trade)
    return trade


async def delete_trade(session: AsyncSession, trade_id: uuid.UUID) -> None:
    trade = await session.get(Trade, trade_id)
    if trade is None:
        raise AppError("trade_not_found", "交易记录不存在", 404)
    await session.delete(trade); await session.commit()


async def list_import_batches(session: AsyncSession, limit: int = 100) -> list[TradeImportBatch]:
    return list((await session.scalars(select(TradeImportBatch).order_by(TradeImportBatch.created_at.desc()).limit(limit))).all())
